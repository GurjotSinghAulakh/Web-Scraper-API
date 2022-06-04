# VERSION 3: Flere lenker

from datetime import datetime  # for å regne tiden til programmet
from tabulate import tabulate  # for å lage tabell i python
from bs4 import BeautifulSoup  # for å gjøre html-koden kompakt
import requests  # for å gjøre request (html-request)

from openpyxl import Workbook, load_workbook

all_brands_array = ["candy", "samsung", "lg"] # må legge til flere brands
all_type_array = ["kombiskap","kjøleskap", "fryser"]

number_of_ads_scraped = 0


# NB: we find and use the first brand_name, in fututre maybe we scrape even more.
def scrape_brand(div_element):
    if div_element is None:
        return "EMPTY"
    else:
        description_text = div_element.text
        description_text_array = description_text.split(" ")
        for word in description_text_array:
            if word.lower() in all_brands_array:
                return word.lower()
        return "Annet"


def scrape_type(div_element):
    if div_element is None:
        return "EMPTY"
    else:
        description_text = div_element.text
        description_text_array = description_text.split(" ")
        for word in description_text_array:
            if word.lower() in all_type_array:
                return word.lower()
        return "Annet"


def scrape(data):
    # variables
    category_title = data.get("category_description")
    category_link = data.get("category_link")
    number_of_ads_to_scrap = int(data.get("number_of_ads"))
    global number_of_ads_scraped
    number_of_ads_scraped = 0

    # defining a work excel book
    wb = Workbook()

    # creating excel sheet
    wb.create_sheet("Gis bort", 0)
    wb.create_sheet("Ønskes kjøpt", 1)
    wb.create_sheet("Til salgs", 2)
    wb.create_sheet("Gi bud", 3)

    ws_gisbort = wb["Gis bort"]
    ws_onskeskjopt = wb["Ønskes kjøpt"]
    ws_tilsalgs = wb["Til salgs"]
    ws_gibud = wb["Gi bud"]

    # start of scraper algorithme
    def funk(page_number):
        if number_of_ads_scraped >= number_of_ads_to_scrap:
            print(f"Total ads from category: {category_title} collected is {number_of_ads_scraped}")
            return

        link = category_link + "&page=" + str(page_number)
        print(link)

        html_code = requests.get(link).text      # extracting the html code from website
        soup = BeautifulSoup(html_code, 'lxml')  # making the html-code compact

        all_ads_on_site = soup.find_all('article', class_="ads__unit")  # finding all ads in the category
        number_of_ads_on_site = len(all_ads_on_site)

        count_ads_on_site = 0
        for ad in all_ads_on_site:
            ad_link_code = ad.find('a', href=True)
            ad_link = ad_link_code['href']

            ad_html_code = requests.get(f'{ad_link}').text  # fetching the html code for each ad
            soup = BeautifulSoup(ad_html_code, 'lxml')  # making the html code compact

            # each ad inn "finn.no" has a section with class_name "panel u-mb16"
            # section has information about ad-title, ad-payment-type and ad-price
            section = soup.find('section', class_="panel u-mb16")

            ad_title = (section.find('h1', class_="u-t2 u-mt16")).text
            ad_payment_type = (section.find('div', class_="u-t4")).text
            ad_price = section.find('div', class_="u-t1")

            # mal: https://www.finn.no/bap/forsale/ad.html?finnkode=258968174
            # finding additional data about the ad
            table_additional_info = soup.find('table', _class="u-width-auto u-mt16")
            ad_info_text = soup.find('div', _class="preserve-linebreaks")      # må gjøre sjekk på om det ikke er tom

            # finding product brand and updating
            product_brand = ""
            product_type = ""
            found_brand = False
            found_type = False

            if table_additional_info is not None:
                # table_th = (table_additional_info.find_all('th', _class="u-text-left u-no-break u-pa0"))
                table_td = (table_additional_info.find_all('td', _class="u-pl16"))
                for td in table_td:
                    if td.text.lower() in all_brands_array:
                        product_brand = td.text.lower()
                        found_brand = True

                    if td.text.lower() in all_type_array:
                        product_type = td.text.lower()
                        found_type = True

                if found_brand is False :
                    product_brand = scrape_brand(ad_info_text)
                if found_type is False :
                    product_type = scrape_type(ad_info_text)
            else:
                product_type = scrape_type(ad_info_text)
                product_brand = scrape_brand(ad_info_text)






            # ad_title = ad_title.text
            # sorting based on payment_type
            if ad_payment_type == "Til salgs":
                if ad_price is None:
                    ws_gibud.append([ad_title, "Gi bud"])
                else:
                    ws_tilsalgs.append([ad_title, ad_price.text])
            elif ad_payment_type == "Ønskes kjøpt":
                ws_onskeskjopt.append([ad_title, "Ønskes kjøpt"])
            else:
                ws_gisbort.append([ad_title, "Gis bort"])

            count_ads_on_site += 1

        # checking if all ads on site has been scraped
        if count_ads_on_site == number_of_ads_on_site:
            number_of_ads_scraped += count_ads_on_site
            print(f"Page {page_number} of category {category_title} is done")
            page_number += 1
            funk(page_number)

    # running the program
    funk(1)
    wb.save(category_title + ".xlsx")


while True:
    category_link = input("category link:")
    if category_link.lower() == "quit":
        break
    description = input("description: ")
    number_of_ads_to_scrap = input("number of ads: ")

    # making dictionary
    data = {"category_link": category_link, "category_description": description, "number_of_ads": number_of_ads_to_scrap}

    scrape(data)

