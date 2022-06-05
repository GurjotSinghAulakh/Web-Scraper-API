'''
    In this version we will try and seperate each uppercategory
    as fridges, sofas and such in their own excel files, as well
    in those excel files, seperate each model/brand in their own sheet
'''

from bs4 import BeautifulSoup                 # to make the html code compact
import requests                               # to make request (html-request)
from openpyxl import Workbook, load_workbook  # To create excel sheets

# TODO: legge til resten av data
# Creating brand arrays:
fridge_brand = ["candy", "samsung", "lg", "whirlpool", "aeg",
                    "husqvarna", "electrolux", "kenwood"]
firdge_type = ["kombiskap", "kjøleskap", "fryser"]



# Dictonary containing information about each product we can scrape
# TODO : Add more products (objects)
dictionary = [
    {
        "category" : "kjøleskap",
        "link" : "https://www.finn.no/bap/forsale/search.html?abTestKey=suggestions&product_category=2.93.3907.292&sort=PUBLISHED_DESC",
        "brand" : fridge_brand,
        "type" : firdge_type

    },
    {   "category" : "komfyr",
         "link" : "https://www.finn.no/bap/forsale/search.html?abTestKey=suggestions&product_category=2.93.3907.73&sort=PUBLISHED_DESC"
    }
]

def ask():
    category_to_be_scraped = input("skriv inn kategori: ").lower()
    # TODO: after solving the problem for max ads to be scraped, we will add this functionality
    number_of_ads = input("Hvor mange annonser ønsker du å scarpe? ")

    for dictionary_element in dictionary:
        if category_to_be_scraped in dictionary_element["category"]:
            category = category_to_be_scraped
            link = dictionary_element["link"]
            brand = dictionary_element["brand"]
            type = dictionary_element["type"]

            scrape(category, link, brand, type, number_of_ads)
            break


# common function
def start():
    print("Hey, Welcome to the scraper!")
    count = 0

    while True:
        if count == 0:
            ask()
            count += 1
        else:
            countinue_to_scrape = input("fortsette ? (y/n) ")
            if countinue_to_scrape.strip().lower() == "n":
                break
            ask()
            count += 1


# NB: we find and use the only first brand_name, in future maybe we scrape even more...
def scrape_brand_from_add_description(div_element, brand_array):
    if div_element is None:
        return "EMPTY"
    else:
        description_text = div_element.text
        description_text_array = description_text.split(" ")
        for word in description_text_array:
            if word.lower() in brand_array:
                return word.lower()
        return "Annet"


# NB: we find and use the only first type name, in future maybe we scrape even more...
def scrape_type_from_add_description(div_element, type_array):
    if div_element is None:
        return "EMPTY"
    else:
        description_text = div_element.text
        description_text_array = description_text.split(" ")
        for word in description_text_array:
            if word.lower() in type_array:
                return word.lower()
        return "Annet"


def scrape(category_title, category_link, brand_array, type_array, number_of_ads_to_scrap):
    # variables
    number_of_ads_scraped = 0
    number_of_ads_to_scrap = int(number_of_ads_to_scrap)

    # defining a work excel book
    wb = Workbook()

    for brand_element in brand_array:
        wb.create_sheet(brand_element)
    wb.create_sheet("Annet merke")


    # start of scraper algorithm
    page_number = 1

    while True:
        if number_of_ads_scraped >= number_of_ads_to_scrap:
            print(f"Total ads from category: {category_title} collected is {number_of_ads_scraped}")
            wb.save(category_title + ".xlsx")
            return

        page_link = category_link + "&page=" + str(page_number)

        page_html_code = requests.get(page_link).text  # extracting the html code from website
        soup = BeautifulSoup(page_html_code, 'lxml')  # making the html-code compact

        all_ads_on_site = soup.find_all('article', class_="ads__unit")  # finding all ads in the category


        for ad in all_ads_on_site:
            ad_link_code = ad.find('a', href=True)
            ad_link = ad_link_code['href']

            # checking if there are any sponsored ads on this category/site
            sponsored_ad = ad.find('span', class_="status status--sponsored u-mb8")
            if sponsored_ad is not None:
                ad_link = "https://www.finn.no" + ad_link

            ad_html_code = requests.get(f'{ad_link}').text  # fetching the html code for each ad
            soup = BeautifulSoup(ad_html_code, 'lxml')  # making the html code compact

            # each ad inn "finn.no" has a section with class_name "panel u-mb16"
            # section has information about ad-title, ad-payment-type and ad-price
            section = soup.find('section', class_="panel u-mb16")

            ad_title = (section.find('h1', class_="u-t2 u-mt16")).text
            ad_payment_type = (section.find('div', class_="u-t4")).text
            ad_price = section.find('div', class_="u-t1")


            # mal: https://www.finn.no/bap/forsale/ad.html?finnkode=258968174
            # finding additional data about the ad
            table_additional_info_html_code = soup.find('table', class_="u-width-auto u-mt16")
            ad_info_text_html_code = soup.find('div', class_="preserve-linebreaks")

            # TODO LIST:
            # must check if it is not empty

            # finding product brand and updating
            product_brand = ""
            product_type = ""
            found_brand = False
            found_type = False

            ad_title_split = ad_title.split(" ")
            for word in ad_title_split:
                if word.lower() in brand_array:
                    product_brand = word.lower()
                    found_brand = True

                if word.lower() in type_array:
                    product_type = word.lower()
                    found_type = True

            if found_brand is False or found_type is False:
                if table_additional_info_html_code is not None:
                    # table_th = (table_additional_info_html_code.find_all('th', class_="u-text-left u-no-break u-pa0"))
                    table_td = (table_additional_info_html_code.find_all('td', class_="u-pl16"))
                    for td in table_td:
                        if td.text.lower() in brand_array:
                            product_brand = td.text.lower()
                            found_brand = True

                        if td.text.lower() in type_array:
                            product_type = td.text.lower()
                            found_type = True

                    if found_brand is False:
                        product_brand = scrape_brand_from_add_description(ad_info_text_html_code, brand_array)
                    if found_type is False:
                        product_type = scrape_type_from_add_description(ad_info_text_html_code, type_array)
                else:
                    product_type = scrape_type_from_add_description(ad_info_text_html_code, type_array)
                    product_brand = scrape_brand_from_add_description(ad_info_text_html_code, brand_array)
            #-----------------------------------

            # Checking if ad_product has is "til salgs", "ønskes kjøpt" or "gis bort"
            # Just the products with a price will be added to the excel-sheet, the rest will be !!ignored!!,
                ## ?? shall we do this or just add them to a seperate sheet?

            if ad_payment_type.lower() == "til salgs":
                if ad_price is None:
                    print("denne er none")
                    print(ad_price)
                    print(ad_link)
                else:
                    price = ad_price.text.replace(" ", "").split("kr")[0]
                    # Sorting ads to differend excel-sheets based on their brand
                    if product_brand.lower() == "annet":
                        ws = wb["Annet merke"]
                        ws.append([ad_title, product_type, price])
                    else:
                        ws = wb[product_brand]
                        ws.append([ad_title, product_type, price])

                number_of_ads_scraped += 1

        # Next-page
        print(f"Page {page_number} of category {category_title} is done")
        page_number += 1


start()