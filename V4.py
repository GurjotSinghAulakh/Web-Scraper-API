# VERSION 4:
# scraped brand and type of the products

# TODO : library tabulate and datetime will be deleted.

from datetime import datetime  # to calculate the time of the program
from tabulate import tabulate  # to create table in python
from bs4 import BeautifulSoup  # to make the html code compact
import requests  # to make request (html-request)
from openpyxl import Workbook, load_workbook  # To create excel sheets

# TODO : skal slettes her i fra og vil bli importert i andre scripts med bestemt kategori.
all_brands_array = ["candy", "samsung", "lg", "whirlpool", "aeg", "husqvarna", "electrolux",
                    "kenwood"]                          # need to add more brands
all_type_array = ["kombiskap", "kjøleskap", "fryser"]   # need to add more types

# number_of_ads_scraped = 0  # A variable which counts the number of ads


# NB: we find and use the only first brand_name, in future maybe we scrape even more...
def scrape_brand_from_add_description(div_element):
    if div_element is None:
        return "EMPTY"
    else:
        description_text = div_element.text
        description_text_array = description_text.split(" ")
        for word in description_text_array:
            if word.lower() in all_brands_array:
                return word.lower()
        return "Annet"


# NB: we find and use the only first type name, in future maybe we scrape even more...
def scrape_type_from_add_description(div_element):
    if div_element is None:
        return "EMPTY"
    else:
        description_text = div_element.text
        description_text_array = description_text.split(" ")
        for word in description_text_array:
            if word.lower() in all_type_array:
                return word.lower()
        return "Annet"


def scrape(data):
    # variables
    category_title = data.get("category_description")
    category_link = data.get("category_link")
    number_of_ads_to_scrap = int(data.get("number_of_ads"))
    number_of_ads_scraped = 0

    # defining a work excel book
    wb = Workbook()

    # creating excel sheet
    wb.create_sheet("Gis bort", 0)
    wb.create_sheet("Ønskes kjøpt", 1)
    wb.create_sheet("Til salgs", 2)
    wb.create_sheet("Gi bud", 3)

    ws_gisbort = wb["Gis bort"]
    ws_onskeskjopt = wb["Ønskes kjøpt"]
    ws_tilsalgs = wb["Til salgs"]
    ws_gibud = wb["Gi bud"]

    # start of scraper algorithm
    page_number = 1
    while True:
        if number_of_ads_scraped >= number_of_ads_to_scrap:
            print(f"Total ads from category: {category_title} collected is {number_of_ads_scraped}")
            wb.save(category_title + ".xlsx")
            return

        link = category_link + "&page=" + str(page_number)
        print(link)

        html_code = requests.get(link).text  # extracting the html code from website
        soup = BeautifulSoup(html_code, 'lxml')  # making the html-code compact

        all_ads_on_site = soup.find_all('article', class_="ads__unit")  # finding all ads in the category
        number_of_ads_on_site = len(all_ads_on_site)

        count_ads_on_site = 0

        for ad in all_ads_on_site:
            ad_link_code = ad.find('a', href=True)
            ad_link = ad_link_code['href']

            # checking if there are any sponsored ads on this category/site
            sponsored_ad = ad.find('span', class_="status status--sponsored u-mb8")
            if sponsored_ad is not None:
                ad_link = "https://www.finn.no" + ad_link

            ad_html_code = requests.get(f'{ad_link}').text  # fetching the html code for each ad
            soup = BeautifulSoup(ad_html_code, 'lxml')  # making the html code compact

            # each ad inn "finn.no" has a section with class_name "panel u-mb16"
            # section has information about ad-title, ad-payment-type and ad-price
            section = soup.find('section', class_="panel u-mb16")

            ad_title = (section.find('h1', class_="u-t2 u-mt16")).text
            ad_payment_type = (section.find('div', class_="u-t4")).text
            ad_price = section.find('div', class_="u-t1")

            # mal: https://www.finn.no/bap/forsale/ad.html?finnkode=258968174
            # finding additional data about the ad
            table_additional_info_html_code = soup.find('table', class_="u-width-auto u-mt16")
            ad_info_text_html_code = soup.find('div', class_="preserve-linebreaks")

            # TODO LIST:
            # must check if it is not empty

            # finding product brand and updating
            product_brand = ""
            product_type = ""
            found_brand = False
            found_type = False

            ad_title_split = ad_title.split(" ")
            for word in ad_title_split:
                if word.lower() in all_brands_array:
                    product_brand = word.lower()
                    found_brand = True

                if word.lower() in all_type_array:
                    product_type = word.lower()
                    found_type = True

            if found_brand is False or found_type is False:
                if table_additional_info_html_code is not None:
                    # table_th = (table_additional_info_html_code.find_all('th', class_="u-text-left u-no-break u-pa0"))
                    table_td = (table_additional_info_html_code.find_all('td', class_="u-pl16"))
                    for td in table_td:
                        if td.text.lower() in all_brands_array:
                            product_brand = td.text.lower()
                            found_brand = True

                        if td.text.lower() in all_type_array:
                            product_type = td.text.lower()
                            found_type = True

                    if found_brand is False:
                        product_brand = scrape_brand_from_add_description(ad_info_text_html_code)
                    if found_type is False:
                        product_type = scrape_type_from_add_description(ad_info_text_html_code)
                else:
                    product_type = scrape_type_from_add_description(ad_info_text_html_code)
                    product_brand = scrape_brand_from_add_description(ad_info_text_html_code)

            # Adding information to the sheets
            if ad_payment_type == "Til salgs":
                if ad_price is None:
                    ws_gibud.append([ad_title, product_brand, product_type, "Gi bud"])
                else:
                    ws_tilsalgs.append([ad_title, product_brand, product_type, ad_price.text])
            elif ad_payment_type == "Ønskes kjøpt":
                ws_onskeskjopt.append([ad_title, product_brand, product_type, "Ønskes kjøpt"])
            else:
                ws_gisbort.append([ad_title, product_brand, product_type, "Gis bort"])

            count_ads_on_site += 1

        # checking if all ads on site has been scraped
        if count_ads_on_site == number_of_ads_on_site:
            number_of_ads_scraped += count_ads_on_site
            print(f"Page {page_number} of category {category_title} is done")
            page_number += 1


while True:
    category_link = input("category link:")
    if category_link.lower() == "quit":
        break
    description = input("description: ")
    number_of_ads_to_scrap = input("number of ads: ")

    # making dictionary
    data = {"category_link": category_link, "category_description": description,
            "number_of_ads": number_of_ads_to_scrap}

    scrape(data)
