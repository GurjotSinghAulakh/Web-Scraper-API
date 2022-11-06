from openpyxl import Workbook, load_workbook  # To create excel sheets
from datetime import datetime

# How to use:
# wb1 have to be the recent file
# wb2 have to be the old file

# Script will check for duplicates in two excel files,
# and add the non-duplicate products and its info to another excel sheet named:
filename = "../Sofa_uke37.xlsx"

wb1 = load_workbook("/Users/gurjotsinghaulakh/Github Repositories/Web-Scraper-API/sofa.xlsx")
ws1 = wb1["Sofa"]

wb2 = load_workbook("/Users/gurjotsinghaulakh/Github Repositories/Web-Scraper-API/Sofa_uke35.xlsx")
ws2 = wb2["Sofa"]

wb = Workbook()
wb.create_sheet("Sofa")
ws = wb["Sofa"]
ws.append(["Varenavn", "Kategori", "Pris", "Merke", "Model", "Lokasjon", "Finn kode"])

for i in range(2, 14_000):
    duplicate = False

    cell1 = f'G{i}'
    finn_kode1 = ws1[cell1].value

    for k in range(2, 14_000):
        cell2 = f'H{k}'
        finn_kode2 = ws2[cell2].value

        if finn_kode1 == finn_kode2:
            duplicate = True
            break

        if finn_kode1 is None and finn_kode2 is None:
            wb.save(filename)
            exit(0)

    if duplicate is False:
        ws.append([ws1[f'A{i}'].value, ws1[f'B{i}'].value, ws1[f'C{i}'].value,
                   ws1[f'D{i}'].value, ws1[f'E{i}'].value, ws1[f'F{i}'].value,
                   ws1[f'G{i}'].value, ws1[f'H{i}'].value])

wb.save(filename)
