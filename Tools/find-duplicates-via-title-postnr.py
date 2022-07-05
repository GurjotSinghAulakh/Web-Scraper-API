from openpyxl import Workbook, load_workbook  # To create excel sheets
from datetime import datetime

# How to use:
# wb1 have to be the recent file
# wb2 have to be the old file

# Script will check for duplicates in two excel files,
# and add the non-duplicate products and its info to another excel sheet named:
filename = "../[STATIC] Scrapped Data/Hvitevarer_uke25.xlsx"

wb1 = load_workbook(
    '/Users/mortaza/Downloads/Hvitevarer-uten-duplikat.xlsx')
ws1 = wb1["Hvitevarer"]

wb2 = load_workbook(
    '/Users/mortaza/Downloads/Hvitevarer--2022-06-15.xlsx')
ws2 = wb2["Hvitevarer"]

wb = Workbook()
wb.create_sheet("Hvitevarer")
ws = wb["Hvitevarer"]
ws.append(["Varenavn", "Under kategori", "Kategori (type)", "Pris", "Merke", "Postnummer", "Lokasjon", "Finn kode"])

for i in range(2, 10_000):
    duplicate = False

    title1_cell = f'A{i}'
    postnr1_cell = f'F{i}'

    title1 = ws1[title1_cell].value
    postnr1 = ws1[postnr1_cell].value

    for k in range(2, 10_000):
        title2_cell = f'A{k}'
        postnr2_cell = f'F{k}'

        title2 = ws2[title2_cell].value
        postnr2 = ws2[postnr2_cell].value

        # checking for duplicates, if ad title and postnr is the same, it means it is
        # most likely a old (duplicate) ad, this will be not be added to the sheet
        if title1 == title2 and postnr1 == postnr2:
            duplicate = True
            break

    if duplicate is False:
        ws.append([ws1[f'A{i}'].value, ws1[f'B{i}'].value, ws1[f'C{i}'].value,
                   ws1[f'D{i}'].value, ws1[f'E{i}'].value, ws1[f'F{i}'].value,
                   ws1[f'G{i}'].value, ws1[f'H{i}'].value])

wb.save(filename)
