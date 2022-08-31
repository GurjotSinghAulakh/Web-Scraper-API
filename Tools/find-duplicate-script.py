from openpyxl import Workbook, load_workbook  # To create excel sheets
from datetime import datetime

# How to use:
# wb1 have to be the recent file
# wb2 have to be the old file

# Script will check for duplicates in two excel files,
# and add the non-duplicate products and its info to another excel sheet named:
filename = '/Users/gurjotsinghaulakh/Library/CloudStorage/OneDrive-OsloMet/Jobb/Secundo/Web-Scraper-API-Github/[STATIC] Scrapped Data/Hvitevarer_uke28.xlsx'


wb1 = load_workbook(
    '/Users/gurjotsinghaulakh/Library/CloudStorage/OneDrive-OsloMet/Jobb/Secundo/Web-Scraper-API-Github/Hvitevarer_12.7.2022.xlsx')
ws1 = wb1["Hvitevarer"]

wb2 = load_workbook(
    './Users/gurjotsinghaulakh/Library/CloudStorage/OneDrive-OsloMet/Jobb/Secundo/Web-Scraper-API-Github/[STATIC] Scrapped Data/Hvitevarer_uke25.xlsx')
ws2 = wb2["Hvitevarer"]

wb = Workbook()
wb.create_sheet("Hvitevarer")
ws = wb["Hvitevarer"]
ws.append(["Varenavn", "Under kategori", "Kategori (type)", "Pris", "Merke", "Postnummer", "Lokasjon"])

for i in range(2, 10_000):
    duplicate = False
    cell1 = f'H{i}'
    finn_kode1 = ws1[cell1].value

    for k in range(2, 10_000):
        cell2 = f'H{k}'
        finn_kode2 = ws2[cell2].value

        if finn_kode1 == finn_kode2:
            duplicate = True
            break

        if finn_kode1 is None and finn_kode2 is None:
            wb.save(filename)
            exit(0)

    if duplicate is False:
        ws.append([ws1[f'A{i}'].value, ws1[f'B{i}'].value, ws1[f'C{i}'].value,
                   ws1[f'D{i}'].value, ws1[f'E{i}'].value,
                   ws1[f'F{i}'].value, ws1[f'G{i}'].value])

wb.save(filename)
