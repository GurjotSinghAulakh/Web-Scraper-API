from openpyxl import Workbook, load_workbook  # To create excel sheets
from datetime import datetime

# How to use:
# wb1 have to be the recent file
# wb2 have to be the old file

# Script will check for duplicates in two excel files,
# and add the non-duplicate products and its info to another excel sheet named:
filename = "../[STATIC] Scrapped Data/Sofa_uke33.xlsx"

wb1 = load_workbook("/Users/gurjotsinghaulakh/Library/CloudStorage/OneDrive-OsloMet/Jobb/Secundo/Web-Scraper-API-Github/sofa.xlsx")
ws1 = wb1["Sofa"]

wb2 = load_workbook("/Users/gurjotsinghaulakh/Library/CloudStorage/OneDrive-OsloMet/Jobb/Secundo/Web-Scraper-API-Github/[STATIC] Scrapped Data/Sofa_uke32.xlsx")
ws2 = wb2["Sofa"]

wb = Workbook()
wb.create_sheet("Sofa")
ws = wb["Sofa"]
ws.append(["Varenavn", "Kategori", "Pris", "Merke", "Model", "Postnummer", "Lokasjon", "Finn kode"])

for i in range(2, 10_000):
    duplicate = False

    title1_cell = f'A{i}'
    postnr1_cell = f'F{i}'

    title1 = ws1[title1_cell].value
    postnr1 = ws1[postnr1_cell].value

    for k in range(2, 10_000):
        title2_cell = f'A{k}'
        postnr2_cell = f'F{k}'

        title2 = ws2[title2_cell].value
        postnr2 = ws2[postnr2_cell].value

        # checking for duplicates, if ad title and postnr is the same, it means it is
        # most likely a old (duplicate) ad, this will be not be added to the sheet
        if title1 == title2 and postnr1 == postnr2:
            duplicate = True
            break

    if duplicate is False:
        ws.append([ws1[f'A{i}'].value, ws1[f'B{i}'].value, ws1[f'C{i}'].value,
                   ws1[f'D{i}'].value, ws1[f'E{i}'].value, ws1[f'F{i}'].value,
                   ws1[f'G{i}'].value, ws1[f'H{i}'].value])

wb.save(filename)
