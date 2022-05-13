# VERSION 3: Flere lenker

from datetime import datetime  # for å regne tiden til programmet
from tabulate import tabulate  # for å lage tabell i python
from bs4 import BeautifulSoup  # for å gjøre html-koden kompakt
import requests  # for å gjøre request (html-request)

from openpyxl import Workbook, load_workbook



number_of_ads_scraped = 0

def scrape(data):
    # variables
    category_title = data.get("category_description")
    category_link = data.get("category_link")
    number_of_ads_to_scrap = int(data.get("number_of_ads"))
    global number_of_ads_scraped
    number_of_ads_scraped = 0

    # defining a work excel book
    wb = Workbook()

    # creating excel sheet
    wb.create_sheet("Gis bort", 0)
    wb.create_sheet("Ønskes kjøpt", 1)
    wb.create_sheet("Til salgs", 2)
    wb.create_sheet("Gi bud", 3)

    ws_gisbort = wb["Gis bort"]
    ws_onskeskjopt = wb["Ønskes kjøpt"]
    ws_tilsalgs = wb["Til salgs"]
    ws_gibud = wb["Gi bud"]

    # start of scraper algorithme
    def funk(page_number):
        if number_of_ads_scraped >= number_of_ads_to_scrap:
            print(f"Total ads from category: {category_title} collected is {number_of_ads_scraped}")
            return

        link = category_link + "&page=" + str(page_number)
        print(link)

        html_code = requests.get(link).text      # extracting the html code from website
        soup = BeautifulSoup(html_code, 'lxml')  # making the html-code compact

        all_ads_on_site = soup.find_all('article', class_="ads__unit")  # finding all ads in the category
        number_of_ads_on_site = len(all_ads_on_site)

        count_ads_on_site = 0
        for ad in all_ads_on_site:
            ad_link_code = ad.find('a', href=True)
            ad_link = ad_link_code['href']

            ad_html_code = requests.get(f'{ad_link}').text  # fetching the html code for each ad
            soup = BeautifulSoup(ad_html_code, 'lxml')  # making the html code compact

            # each ad inn "finn.no" has a section with class_name "panel u-mb16"
            # section has information about ad-title, ad-payment-type and ad-price
            section = soup.find('section', class_="panel u-mb16")

            ad_title = (section.find('h1', class_="u-t2 u-mt16")).text
            ad_payment_type = (section.find('div', class_="u-t4")).text
            ad_price = section.find('div', class_="u-t1")

            # ad_title = ad_title.text
            # sorting based on payment_type
            if ad_payment_type == "Til salgs":
                if ad_price is None:
                    ws_gibud.append([ad_title, "Gi bud"])
                else:
                    ws_tilsalgs.append([ad_title, ad_price.text])
            elif ad_payment_type == "Ønskes kjøpt":
                ws_onskeskjopt.append([ad_title, "Ønskes kjøpt"])
            else:
                ws_gisbort.append([ad_title, "Gis bort"])

            count_ads_on_site += 1

        # checking if all ads on site has been scraped
        if count_ads_on_site == number_of_ads_on_site:
            number_of_ads_scraped += count_ads_on_site
            print(f"Page {page_number} of category {category_title} is done")
            page_number += 1
            funk(page_number)

    # running the program
    funk(1)
    wb.save(category_title + ".xlsx")


while True:
    category_link = input("category link:")
    if category_link.lower() == "quit":
        break
    description = input("description: ")
    number_of_ads_to_scrap = input("number of ads: ")

    # making dictionary
    data = {"category_link": category_link, "category_description": description, "number_of_ads": number_of_ads_to_scrap}

    scrape(data)

