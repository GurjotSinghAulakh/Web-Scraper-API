# VERSION 2: all data legges i hver sin sheet

from datetime import datetime  # for å regne tiden til programmet
from tabulate import tabulate  # for å lage tabell i python
from bs4 import BeautifulSoup  # for å gjøre html-koden kompakt
import requests  # for å gjøre request (html-request)

from openpyxl import Workbook, load_workbook

wb = Workbook()
wb.create_sheet("Gis bort", 0)
wb.create_sheet("Ønskes kjøpt", 1)
wb.create_sheet("Til salgs", 2)
wb.create_sheet("Gi bud", 3)

ws_gisbort = wb["Gis bort"]
ws_onskeskjopt = wb["Ønskes kjøpt"]
ws_tilsalgs = wb["Til salgs"]
ws_gibud = wb["Gi bud"]

ws_gisbort.append(["Produkt tittel", "Produkt pris"])
ws_onskeskjopt.append(["Produkt tittel", "Produkt pris"])
ws_tilsalgs.append(["Produkt tittel", "Produkt pris"])
ws_gibud.append(["Produkt tittel", "Produkt pris"])

start_time = datetime.now()  # starter tid
table = [['TITLE', 'PRICE']]

total_ads_collected = 0


def funk(site, page, total_ads_that_we_want_to_collect_data_from):
    global total_ads_collected
    if total_ads_collected >= total_ads_that_we_want_to_collect_data_from:
        print(f"Total ads collected data from {total_ads_collected}")
        return

    link = site + "&page=" + str(page)
    print(link)

    html_text = requests.get(link).text  # extracting the html code from website
    soup = BeautifulSoup(html_text, 'lxml')  # making the html-code compact

    all_ads_on_site = soup.find_all('article', class_="ads__unit")  # finding all ads in the category
    total_ads_on_site = len(all_ads_on_site)

    # # the next button
    # next_button = soup.find('a', class_="button button--pill button--has-icon button--icon-right")
    # # next_button_link = next_button['href']
    #
    # if next_button is None:
    #     print("Du har kommet til slutten av søket!")
    #     return

    count = 0
    for ad in all_ads_on_site:
        product_link_code = ad.find('a', href=True)
        ad_link = product_link_code['href']

        ad_html_code = requests.get(f'{ad_link}').text  # fetching the html code for each ad
        soup = BeautifulSoup(ad_html_code, 'lxml')  # making the html code compact

        # each ad inn "finn.no" has a section with class_name "panel u-mb16"
        section = soup.find('section', class_="panel u-mb16")

        # print(section.text)

        titles = section.find('h1', class_="u-t2 u-mt16")
        prices_tilSalgs = section.find('div', class_="u-t1")
        payment_type = section.find('div', class_="u-t4").text

        final_product_price = 0

        if payment_type == "Til salgs":
            if prices_tilSalgs is None:
                final_product_price = "Gi bud"
                ws_gibud.append([titles.text, final_product_price])
            else:
                final_product_price = prices_tilSalgs.text
                ws_tilsalgs.append([titles.text, final_product_price])
        elif payment_type == "Ønskes kjøpt":
            final_product_price = "Ønskes kjøpt"
            ws_onskeskjopt.append([titles.text, final_product_price])
        else:
            final_product_price = "Gis bort"
            ws_gisbort.append([titles.text, final_product_price])

        table.append([titles.text, final_product_price])
        count += 1

    if count == total_ads_on_site:
        print(f"page {page} is done")
        total_ads_collected += total_ads_on_site
        page += 1
        funk(site, page, total_ads_that_we_want_to_collect_data_from)


# funk(link, side, antall produkter vi ønsker å hente data fra)
funk('https://www.finn.no/bap/forsale/search.html?abTestKey=suggestions&sort=RELEVANCE&sub_category=1.93.3907', 1, 2000)
end_time = datetime.now()
print(tabulate(table))
print(end_time - start_time)
wb.save("test.xlsx")
