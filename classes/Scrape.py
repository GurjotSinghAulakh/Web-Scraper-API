import threading                # to enable threading

import requests                 # to make request (html-request)
from bs4 import BeautifulSoup   # to make the html code compact
from openpyxl import Workbook, load_workbook  # to create excel sheets

class Scrape:
    @staticmethod
    def scrape():
       pass

    @staticmethod
    def scrape_title(div, array):
        pass

    @staticmethod
    def scrape_desciprion(div, array):
        # handling None-pointer exception
        if div is None:
            print("[Warning] : description element is empty")
        else:
            description_text = div.text
            description_text_array = description_text.split(" ")
            for word in description_text_array:
                if word.lower() in array:
                    return word.lower()
            return ""

    @staticmethod
    def scrape_table(div, array):
        pass

    @staticmethod
    def scrape_finn_code(div):
        pass

    @staticmethod
    def scrape_location(div):
        pass

    @staticmethod
    def find_duplicates(new_file, old_file, category):
        filename = "Hvitevarer_uke25.xlsx"

        wb1 = load_workbook(new_file)
        ws1 = wb1[category]

        wb2 = load_workbook(old_file)
        ws2 = wb2[category]

        wb = Workbook()
        wb.create_sheet(category)
        ws = wb[category]

        if category.lower()== "hvitevarer":
            ws.append(["Varenavn", "Under kategori", "Kategori (type)", "Pris", "Merke", "Postnummer", "Lokasjon"])
            for i in range(2, 10_000):
                duplicate = False
                cell1 = f'H{i}'
                finn_kode1 = ws1[cell1].value

                for k in range(2, 10_000):
                    cell2 = f'H{k}'
                    finn_kode2 = ws2[cell2].value

                    if finn_kode1 == finn_kode2:
                        duplicate = True
                        break

                    if finn_kode1 is None and finn_kode2 is None:
                        wb.save(filename)
                        exit(0)

                if duplicate is False:
                    ws.append([ws1[f'A{i}'].value, ws1[f'B{i}'].value, ws1[f'C{i}'].value,
                               ws1[f'D{i}'].value, ws1[f'E{i}'].value,
                               ws1[f'F{i}'].value, ws1[f'G{i}'].value])

            wb.save(filename)

        elif category == "sofa":
            print(category)