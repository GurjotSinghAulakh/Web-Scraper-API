# from openpyxl import Workbook, load_workbook  # To create excel sheets
#
# wb = load_workbook('/Users/mortaza/Downloads/2022-06-08.xlsx')
# ws = wb["Hvitevarer"]
#
# andre_hvitevarer = 0
# frysere = 0
# innbyggingsovner = 0
# kjøleskap = 0
# komfyrer = 0
# mikrobølgeovner = 0
# oppvaskmaskiner = 0
# platetopper = 0
# tørketromler = 0
# vaskemaskiner = 0
# ventilatorer = 0
#
# for i in range(2, 9101):
#     cell = f'B{i}'
#     verdi = ws[cell].value
#     if verdi == "andre hvitevarer":
#         andre_hvitevarer += 1
#     elif verdi == "frysere":
#         frysere += 1
#     elif verdi == "innbyggingsovner":
#         innbyggingsovner += 1
#     elif verdi == "kjøleskap":
#         kjøleskap += 1
#     elif verdi == "komfyrer":
#         komfyrer += 1
#     elif verdi == "mikrobølgeovner":
#         mikrobølgeovner += 1
#     elif verdi == "oppvaskmaskiner":
#         oppvaskmaskiner += 1
#     elif verdi == "platetopper":
#         platetopper += 1
#     elif verdi == "tørketromler":
#         tørketromler += 1
#     elif verdi == "vaskemaskiner":
#         vaskemaskiner += 1
#     elif verdi == "ventilatorer":
#         ventilatorer += 1
#     else:
#         print(f"HVA FAEN {verdi}")
#
# print(andre_hvitevarer)
# print(frysere)
# print(innbyggingsovner)
# print(kjøleskap)
# print(komfyrer)
# print(mikrobølgeovner)
# print(oppvaskmaskiner)
# print(platetopper)
# print(tørketromler)
# print(vaskemaskiner)
# print(ventilatorer)
#
from openpyxl import Workbook, load_workbook  # To create excel sheets
wb1 = load_workbook('/Users/mortaza/Downloads/2022-06-08.xlsx')
ws1 = wb1["Hvitevarer"]

wb2 = load_workbook('/Users/mortaza/Downloads/2022-06-08.xlsx')
ws2 = wb2["Hvitevarer"]

for i in range(2, 9200):
    cell1 = f'A{i}'
    verdi1 = ws1[cell1].value
    for b in range(2, 9200):
        cell2 = f'A{b}'
        verdi2 = ws2[cell2].value

        if verdi1 == verdi2:
            print(f"Dette er duplikat {verdi2}")
